import openpyxl

file = "G:\HDFC Credit Card\Vijay_Data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active
print(sheet)
sheet.cell(1,1).value="S.NO"
sheet.cell(1,2).value="Names"
sheet.cell(1,3).value="math"
sheet.cell(1,4).value="Physics"
workbook.save(file)

sheet.cell(2,1).value=1
sheet.cell(3,1).value=2
sheet.cell(4,1).value=3
sheet.cell(5,1).value=4
workbook.save(file)
