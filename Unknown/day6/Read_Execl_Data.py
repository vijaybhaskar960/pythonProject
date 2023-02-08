import openpyxl


file = "G:\HDFC Credit Card\EMI Bills.xlsx"
workbook = openpyxl.load_workbook(file)
Sheet = workbook["Sheet1"]
rows = Sheet.max_row
cols = Sheet.max_column
print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(Sheet.cell(r,c).value,end="     ")
    print()

