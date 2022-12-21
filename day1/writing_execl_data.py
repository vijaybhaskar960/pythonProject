import openpyxl

file = "G:\HDFC Credit Card\Book1.xlsx"
workbbok = openpyxl.load_workbook(file)
sheet = workbbok.active     # get active sheet from execl
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="welcome"
#
# workbbok.save(file)

sheet.cell(1,1).value=123
sheet.cell(1,2).value='smit'
sheet.cell(1,3).value="Engg"

sheet.cell(2,1).value=56
sheet.cell(2,2).value="sai"
sheet.cell(2,3).value="Developer"

sheet.cell(3,1).value=897
sheet.cell(3,2).value="sheet"
sheet.cell(3,3).value=" accosite"

workbbok.save(file)


