import openpyxl

file = "F:\HDFC Credit Card\Book1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active

sheet.cell(4,1).value = "Vaishu"
sheet.cell(4,2).value = "IT Job"
sheet.cell(4,3).value = "Salary"
sheet.cell(4,4).value = "Software"

workbook.save(file)
